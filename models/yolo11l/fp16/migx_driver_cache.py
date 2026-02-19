import os
import sys
import re
import subprocess
import platform
from time import perf_counter

model_source_name = 'yolov11l_fp16{batch}b.onnx'
migx_binary = 'migraphx-driver.exe' if platform.system() == 'Windows' else 'migraphx-driver'

# Setting batch sizes from command line
batches = [1]
if '--batch-size' in sys.argv:
  try:
    batches = [int(x) for x in sys.argv[sys.argv.index('--batch-size') + 1].split(',')]
  except Exception as e:
    print(f'{{ "Error": "Failed to set batch size {e}, using default [{", ".join(map(str, batches))}]" }},')

if not os.path.exists('./temp'):
  os.makedirs('./temp')

# Get migraphx-driver version
def get_migraphx_version():
  global migx_binary
  try:
    result = subprocess.run([migx_binary, '-v'], capture_output=True, text=True, timeout=10)
    version_output = result.stdout.strip() + result.stderr.strip()
    # Extract version from output like "[ MIGraphX Version: 2.16.0.20250912-17-879-gef642c96e ]"
    match = re.search(r'MIGraphX Version:\s*([\d\.\-\w]+)', version_output)
    if match:
      return match.group(1)
    return version_output
  except Exception as e:
    return f"Unknown (Error: {e})"

migraphx_version = get_migraphx_version()
print(f'{{ "MIGraphX Version": "{migraphx_version}" }},')

def parse_migraphx_output(output):
  """Parse migraphx-driver perf output and extract performance metrics"""
  data = {}
  
  # Extract batch size
  batch_match = re.search(r'Batch size:\s*(\d+)', output)
  if batch_match:
    data['batch_size'] = int(batch_match.group(1))
  
  # Extract rate (inferences/sec)
  rate_match = re.search(r'Rate:\s*([\d\.]+)\s*inferences/sec', output)
  if rate_match:
    data['rate'] = float(rate_match.group(1))
  
  # Extract total time with stats
  total_time_match = re.search(
    r'Total time:\s*([\d\.]+)ms\s*\(Min:\s*([\d\.]+)ms,\s*Max:\s*([\d\.]+)ms,\s*Mean:\s*([\d\.]+)ms,\s*Median:\s*([\d\.]+)ms\)',
    output
  )
  if total_time_match:
    data['total_time'] = float(total_time_match.group(1)) / 1000  # Convert to seconds
    data['min_time'] = float(total_time_match.group(2)) / 1000
    data['max_time'] = float(total_time_match.group(3)) / 1000
    data['mean_time'] = float(total_time_match.group(4)) / 1000
    data['median_time'] = float(total_time_match.group(5)) / 1000
  
  # Extract percentiles
  percentiles_match = re.search(
    r'Percentiles\s*\(90%,\s*95%,\s*99%\):\s*\(([\d\.]+)ms,\s*([\d\.]+)ms,\s*([\d\.]+)ms\)',
    output
  )
  if percentiles_match:
    data['p90'] = float(percentiles_match.group(1)) / 1000
    data['p95'] = float(percentiles_match.group(2)) / 1000
    data['p99'] = float(percentiles_match.group(3)) / 1000
  
  # Extract instructions time
  instructions_match = re.search(r'Total instructions time:\s*([\d\.]+)ms', output)
  if instructions_match:
    data['instructions_time'] = float(instructions_match.group(1)) / 1000
  
  # Extract overhead
  overhead_match = re.search(r'Overhead time:\s*([\d\.\-]+)ms,\s*([\d\.\-]+)ms', output)
  if overhead_match:
    data['overhead_time1'] = float(overhead_match.group(1)) / 1000
    data['overhead_time2'] = float(overhead_match.group(2)) / 1000
  
  return data

# Import try_export_model from yolo11l common
from .common import try_export_model

inference_times = {}
compile_times = {}
all_results = {}

for batch in batches:
  print(f'{{ "Processing Batch": {batch} }},')
  
  model_name = model_source_name.format(batch=batch)
  model_path = os.path.join('./temp', model_name)
  mxr_name = model_name[:-4] + 'mxr'
  mxr_path = os.path.join('./temp', mxr_name)
  
  # Step 1: Check if model exists, if not export it
  if not os.path.exists(model_path):
    print(f'{{ "Exporting Model": "{model_path}" }},')
    try:
      try_export_model(model_path, batch, half_precision=False)
    except Exception as e:
      print(f'{{ "Error": "Failed to export model {e}" }},')
      continue
  
  # Step 2: Check if .mxr exists, if not compile it
  if not os.path.exists(mxr_path):
    print(f'{{ "Compiling MXR": "{mxr_path}" }},')
    try:
      start_time = perf_counter()
      compile_cmd = [migx_binary, 'compile', model_path, '--gpu', '--enable-offload-copy', '--binary', '-o', mxr_path]
      result = subprocess.run(compile_cmd, capture_output=True, text=True)
      compile_time = perf_counter() - start_time
      
      if result.returncode != 0:
        print(f'{{ "Error": "Failed to compile model: {result.stderr}" }},')
        continue
      
      print(f'{{ "Compile Time": {compile_time} }},')
      compile_times[batch] = compile_time
    except Exception as e:
      print(f'{{ "Error": "Failed to compile model {e}" }},')
      continue
  else:
    compile_times[batch] = 0
  
  # Step 3: Run perf command
  print(f'{{ "Running Performance Test": "{mxr_path}" }},')
  try:
    perf_cmd = [migx_binary, 'perf', '--enable-offload-copy', '--migraphx', mxr_path]
    result = subprocess.run(perf_cmd, capture_output=True, text=True)
    
    if result.returncode != 0:
      print(f'{{ "Error": "Failed to run perf: {result.stderr}" }},')
      continue
    
    # Parse output
    perf_data = parse_migraphx_output(result.stdout + result.stderr)
    all_results[batch] = perf_data
    
    print(f'{{ "Performance Data": {perf_data} }},')
    
    # Create inference times list for report generation
    # MIGraphX doesn't provide individual run times, so we'll use the mean
    if 'mean_time' in perf_data:
      # Simulate multiple runs with the mean (for report compatibility)
      num_runs = 10
      inference_times[batch] = [perf_data['mean_time']] * num_runs
      inference_times[batch].append({
        "Minimum": perf_data.get('min_time', perf_data['mean_time']),
        "Maximum": perf_data.get('max_time', perf_data['mean_time']),
        "Average": perf_data['mean_time']
      })
    
  except Exception as e:
    print(f'{{ "Error": "Failed to run performance test {e}" }},')
    continue

# Step 4 & 5: Generate report with openpyxl
if inference_times:
  print('{ "Generating Report": "Starting" },')
  try:
    import openpyxl
    from openpyxl.chart import LineChart, Reference, Series
    from openpyxl.chart.layout import Layout, ManualLayout
    from openpyxl.utils import get_column_letter
    from copy import deepcopy
    import datetime
    
    wb = openpyxl.Workbook()
    main_sheet = wb.active
    main_sheet.title = "Overview"
    inference_sheet = wb.create_sheet("Inference")
    
    # Build inference sheet
    inference_sheet.append(["Inference times (MIGraphX)"])
    inference_sheet.append(["Metric"] + [f"Batch {batch}" for batch in batches])
    
    offset_col = 2
    offset_stat_row = inference_sheet.max_row + 1
    
    inference_sheet.append(["Average"])
    inference_sheet.append(["Median"])
    inference_sheet.append(["90th Percentile"])
    inference_sheet.append(["95th Percentile"])
    inference_sheet.append(["99th Percentile"])
    inference_sheet.append(["Minimum"])
    inference_sheet.append(["Maximum"])
    inference_sheet.append(["IPS (Average)"])
    inference_sheet.append(["IPS (Median)"])
    inference_sheet.append(["IPS (90th Percentile)"])
    inference_sheet.append(["IPS (95th Percentile)"])
    inference_sheet.append(["IPS (99th Percentile)"])
    inference_sheet.append(["BPS (Average)"])
    inference_sheet.append(["BPS (Median)"])
    inference_sheet.append(["BPS (90th Percentile)"])
    inference_sheet.append(["BPS (95th Percentile)"])
    inference_sheet.append(["BPS (99th Percentile)"])
    inference_sheet.append(["Compile Time"])
    
    # Fill in data
    for batch_index, batch in enumerate(batches):
      col_letter = get_column_letter(offset_col + batch_index)
      if batch in all_results:
        data = all_results[batch]
        inference_sheet[col_letter + str(offset_stat_row + 0)] = data.get('mean_time', 0)
        inference_sheet[col_letter + str(offset_stat_row + 1)] = data.get('median_time', 0)
        inference_sheet[col_letter + str(offset_stat_row + 2)] = data.get('p90', 0)
        inference_sheet[col_letter + str(offset_stat_row + 3)] = data.get('p95', 0)
        inference_sheet[col_letter + str(offset_stat_row + 4)] = data.get('p99', 0)
        inference_sheet[col_letter + str(offset_stat_row + 5)] = data.get('min_time', 0)
        inference_sheet[col_letter + str(offset_stat_row + 6)] = data.get('max_time', 0)
        
        # Calculate IPS (Inferences Per Second)
        if data.get('mean_time', 0) > 0:
          inference_sheet[col_letter + str(offset_stat_row + 7)] = 1 / data['mean_time']
        if data.get('median_time', 0) > 0:
          inference_sheet[col_letter + str(offset_stat_row + 8)] = 1 / data['median_time']
        if data.get('p90', 0) > 0:
          inference_sheet[col_letter + str(offset_stat_row + 9)] = 1 / data['p90']
        if data.get('p95', 0) > 0:
          inference_sheet[col_letter + str(offset_stat_row + 10)] = 1 / data['p95']
        if data.get('p99', 0) > 0:
          inference_sheet[col_letter + str(offset_stat_row + 11)] = 1 / data['p99']
        
        # Calculate BPS (Batches Per Second) = IPS * batch_size
        inference_sheet[col_letter + str(offset_stat_row + 12)] = f"={batch} * {col_letter + str(offset_stat_row + 7)}"
        inference_sheet[col_letter + str(offset_stat_row + 13)] = f"={batch} * {col_letter + str(offset_stat_row + 8)}"
        inference_sheet[col_letter + str(offset_stat_row + 14)] = f"={batch} * {col_letter + str(offset_stat_row + 9)}"
        inference_sheet[col_letter + str(offset_stat_row + 15)] = f"={batch} * {col_letter + str(offset_stat_row + 10)}"
        inference_sheet[col_letter + str(offset_stat_row + 16)] = f"={batch} * {col_letter + str(offset_stat_row + 11)}"
        
        inference_sheet[col_letter + str(offset_stat_row + 17)] = compile_times.get(batch, 0)
    
    # Create charts
    # Metrics chart
    chart = LineChart()
    chart.title = "Metrics (Time)"
    chart.x_axis.title = "Batch Size"
    chart.y_axis.title = "Time (s)"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    metrics = ["Average", "Median", "90th Percentile", "95th Percentile", "99th Percentile", "Minimum", "Maximum"]
    for metric_index in range(len(metrics)):
      series = Series(
        values=Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row + metric_index, 
                        max_col=offset_col + len(batches) - 1, max_row=offset_stat_row + metric_index),
        title=f"{metrics[metric_index]}"
      )
      series.marker.symbol = "circle"
      series.marker.size = 6
      chart.series.append(series)
    batch_titles = Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row - 1, 
                            max_col=offset_col + len(batches) - 1, max_row=offset_stat_row - 1)
    chart.set_categories(batch_titles)
    chart.legend.position = 'b'
    chart.layout = Layout(manualLayout=ManualLayout(x=0.02, y=0.02, h=0.65, w=0.9))
    chart.width = 25
    inference_sheet.add_chart(chart, get_column_letter(len(batches) + 2) + "1")
    main_sheet.add_chart(deepcopy(chart), "F5")
    
    # IPS chart
    chart = LineChart()
    chart.title = "IPS (Inferences Per Second)"
    chart.x_axis.title = "Batch Size"
    chart.y_axis.title = "Inferences Per Second"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    metrics = ["Average", "Median", "90th Percentile", "95th Percentile", "99th Percentile"]
    for metric_index in range(len(metrics)):
      series = Series(
        values=Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row + 7 + metric_index,
                        max_col=offset_col + len(batches) - 1, max_row=offset_stat_row + 7 + metric_index),
        title=f"{metrics[metric_index]}"
      )
      series.marker.symbol = "circle"
      series.marker.size = 6
      chart.series.append(series)
    chart.set_categories(batch_titles)
    chart.legend.position = 'b'
    chart.layout = Layout(manualLayout=ManualLayout(x=0.02, y=0.02, h=0.65, w=0.9))
    chart.width = 15
    inference_sheet.add_chart(chart, get_column_letter(len(batches) + 2) + "16")
    main_sheet.add_chart(deepcopy(chart), "F20")
    
    # BPS chart
    chart = LineChart()
    chart.title = "BPS (Batches Per Second)"
    chart.x_axis.title = "Batch Size"
    chart.y_axis.title = "Batches Per Second"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    for metric_index in range(len(metrics)):
      series = Series(
        values=Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row + 12 + metric_index,
                        max_col=offset_col + len(batches) - 1, max_row=offset_stat_row + 12 + metric_index),
        title=f"{metrics[metric_index]}"
      )
      series.marker.symbol = "circle"
      series.marker.size = 6
      chart.series.append(series)
    chart.set_categories(batch_titles)
    chart.legend.position = 'b'
    chart.layout = Layout(manualLayout=ManualLayout(x=0.02, y=0.02, h=0.65, w=0.9))
    chart.width = 15
    inference_sheet.add_chart(chart, get_column_letter(len(batches) + 11) + "16")
    main_sheet.add_chart(deepcopy(chart), "P20")
    
    # Build overview page
    report_datetime = datetime.datetime.now()
    main_sheet.column_dimensions[get_column_letter(1)].width = 30
    main_sheet.append(['Model:', 'yolov11l (MIGraphX)'])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)
    main_sheet.append(['Description:', 'YOLO11L model running on MIGraphX'])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)
    main_sheet.append(['Run Command:', ' '.join(sys.argv)])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)
    main_sheet.append(['Report Date:', report_datetime.strftime('%Y-%m-%d %H:%M:%S')])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=6)
    main_sheet.append(['Batches:', *batches])
    main_sheet.append([])
    main_sheet.append(['System Information:'])
    try:
      main_sheet.append(['Hostname:', platform.node()])
      main_sheet.append(['OS:', platform.system()])
      main_sheet.append(['OS Version:', platform.version()])
      main_sheet.append(['OS Release:', platform.release()])
    except Exception as e:
      main_sheet.append([f'Cannot get OS information {e}'])
    main_sheet.append(['Python Version:', sys.version])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)
    main_sheet.append(['MIGraphX Version:', migraphx_version])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)
    
    try:
      main_sheet.append(['CPU:', platform.processor()])
      import reports
      accelerators = reports.enumerate_accelerators()
      for item in accelerators['gpu']:
        main_sheet.append(['GPU:', item['name']])
      for item in accelerators['npu']:
        main_sheet.append(['NPU:', item['name']])
    except Exception as e:
      main_sheet.append([f'Cannot get accelerators information {e}'])
    
    try:
      main_sheet.append([])
      main_sheet.append(['Environment Variables:'])
      for key, value in sorted(os.environ.items(), key=lambda x: x[0]):
        main_sheet.append([key, value])
    except Exception as e:
      main_sheet.append([f'Cannot get environment variables {e}'])
    
    # Save workbook
    workbook_name = f"{platform.node().lower()}_models.yolo11l.fp16.migx_driver_cache_{report_datetime.strftime('%Y%m%d_%H%M%S')}.xlsx"
    workbook_path = workbook_name
    
    reports_path = os.path.join(os.path.dirname(__file__), '..', '..', '..', 'reports', report_datetime.strftime("%Y%m%d"))
    if not os.path.exists(reports_path):
      os.makedirs(reports_path)
      try:
        from shutil import copy
        copy(os.path.join(os.path.dirname(__file__), '..', '..', '..', "!StatViewer.xlsm"),
             os.path.join(reports_path, "!StatViewer.xlsm"))
      except Exception as e:
        print(f'{{ "Error": "Failed to copy !StatViewer.xlsm {e}" }}')
    
    wb.save(workbook_path)
    os.rename(workbook_path, os.path.join(reports_path, workbook_path))
    
    print(f'{{ "Workbook": "{os.path.join(reports_path, workbook_path).replace('\\', '/')}" }},')
    
  except Exception as e:
    print(f'{{ "Error": "Failed to generate report {e}" }},')
    import traceback
    traceback.print_exc()

print('{ "Status": "Done" }')

if __name__ != "__main__":
  exit(0)
