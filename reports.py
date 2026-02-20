import sys
import os
import subprocess
import datetime
import platform
from copy import deepcopy
from shutil import copy, which

def performance_report(model,model_name, read_times, inference_times, warm_up_times, batches):
  try:
    import openpyxl
    from openpyxl.chart import LineChart, Reference, Series
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.chart.layout import Layout, ManualLayout
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    main_sheet = wb.active
    read_sheet = wb.create_sheet("Read")
    inference_sheet = wb.create_sheet("Inference")

    read_sheet.append(["Reading times"])

    offset_col = 2
    offset_stat_row = read_sheet.max_row + 1
    read_sheet.append(["Average"])
    read_sheet.append(["Median"])
    read_sheet.append(["90th Percentile"])
    read_sheet.append(["95th Percentile"])
    read_sheet.append(["99th Percentile"])
    read_sheet.append(["Minimum"])
    read_sheet.append(["Maximum"])

    read_sheet.append(["Run", "Time (s)"])

    col_letter = get_column_letter(offset_col)
    offset_row = read_sheet.max_row + 1
    last_row = offset_row + len(read_times) - 2
    read_sheet[col_letter + str(offset_stat_row + 0)] = "=AVERAGE(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"
    read_sheet[col_letter + str(offset_stat_row + 1)] = "=MEDIAN(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"
    read_sheet[col_letter + str(offset_stat_row + 2)] = "=_xlfn.PERCENTILE.INC(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ", 0.9)"
    read_sheet[col_letter + str(offset_stat_row + 3)] = "=_xlfn.PERCENTILE.INC(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ", 0.95)"
    read_sheet[col_letter + str(offset_stat_row + 4)] = "=_xlfn.PERCENTILE.INC(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ", 0.99)"
    read_sheet[col_letter + str(offset_stat_row + 5)] = "=MIN(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"
    read_sheet[col_letter + str(offset_stat_row + 6)] = "=MAX(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"

    offset_col = 2
    offset_row = read_sheet.max_row + 1
    last_row = offset_row + len(read_times) - 2

    idx = 1
    for item in read_times[:-1]:  
        read_sheet.append([idx, item])
        idx += 1

    series = Series(values=Reference(read_sheet, min_col=offset_col, min_row=offset_row, max_col=offset_col, max_row=last_row), title="Reading times")
    if len(read_times) <= 2:
        series.marker.symbol = "circle"
        series.marker.size = 6
    chart = LineChart()
    chart.series.append(series)
    chart.title = "Reading times"
    chart.x_axis.title = "Run"
    chart.y_axis.title = "Time (s)"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.legend = None
    chart.layout=Layout(
        manualLayout=ManualLayout(
            x=0.02, y=0.02,
            h=0.75, w=0.9,
        )
    )
    read_sheet.add_chart(chart, "C1")
    main_sheet.add_chart(deepcopy(chart), "F35")

    inference_sheet.append(["Inference times"])
    inference_table = [[] for _ in range(len(batches))]
    x_axis_max = 0
    x_axis_min = float('inf')
    for batch_index in range(len(batches)):
        for item in inference_times[batches[batch_index]][:-1]:
            inference_table[batch_index].append(item)
            x_axis_max = max(x_axis_max, item)
            x_axis_min = min(x_axis_min, item)

    # Table header
    inference_sheet.append(["Metric"] + [f"Batch {batch}" for batch in batches])
    # Aggregated statistics
    inference_sheet.column_dimensions[get_column_letter(1)].width = 30
    offset_col = 2
    offset_stat_row = inference_sheet.max_row + 1
    inference_sheet.append(["Average"])
    inference_sheet.append(["Median"])
    inference_sheet.append(["90th Percentile"])
    inference_sheet.append(["95th Percentile"])
    inference_sheet.append(["99th Percentile"])
    inference_sheet.append(["Minimum"])
    inference_sheet.append(["Maximum"])
    inference_sheet.append(["IPS (Average)"])
    inference_sheet.append(["IPS (Median)"])
    inference_sheet.append(["IPS (90th Percentile)"])
    inference_sheet.append(["IPS (95th Percentile)"])
    inference_sheet.append(["IPS (99th Percentile)"])
    inference_sheet.append(["BPS (Average)"])
    inference_sheet.append(["BPS (Median)"])
    inference_sheet.append(["BPS (90th Percentile)"])
    inference_sheet.append(["BPS (95th Percentile)"])
    inference_sheet.append(["BPS (99th Percentile)"])
    inference_sheet.append(["Warm Up Time"])

    # Table header
    inference_sheet.append(["Run"] + [f"Batch {batch}" for batch in batches])

    offset_col = 2
    offset_row = inference_sheet.max_row + 1
    last_row = offset_row + len(inference_table[0]) - 1

    for batch_index in range(len(batches)):
        col_letter = get_column_letter(offset_col + batch_index)
        inference_sheet[col_letter + str(offset_stat_row + 0)] = "=AVERAGE(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"
        inference_sheet[col_letter + str(offset_stat_row + 1)] = "=MEDIAN(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"
        inference_sheet[col_letter + str(offset_stat_row + 2)] = "=_xlfn.PERCENTILE.INC(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ", 0.9)"
        inference_sheet[col_letter + str(offset_stat_row + 3)] = "=_xlfn.PERCENTILE.INC(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ", 0.95)"
        inference_sheet[col_letter + str(offset_stat_row + 4)] = "=_xlfn.PERCENTILE.INC(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ", 0.99)"
        inference_sheet[col_letter + str(offset_stat_row + 5)] = "=MIN(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"
        inference_sheet[col_letter + str(offset_stat_row + 6)] = "=MAX(" + col_letter + str(offset_row) + ":" + col_letter + str(last_row) + ")"
        # Inference Per Second depending on calculated time
        inference_sheet[col_letter + str(offset_stat_row + 7)] = "=1 / " + col_letter + str(offset_stat_row + 0)
        inference_sheet[col_letter + str(offset_stat_row + 8)] = "=1 / " + col_letter + str(offset_stat_row + 1)
        inference_sheet[col_letter + str(offset_stat_row + 9)] = "=1 / " + col_letter + str(offset_stat_row + 2)
        inference_sheet[col_letter + str(offset_stat_row + 10)] = "=1 / " + col_letter + str(offset_stat_row + 3)
        inference_sheet[col_letter + str(offset_stat_row + 11)] = "=1 / " + col_letter + str(offset_stat_row + 4)
        # Batch Per Second depending on IPS
        inference_sheet[col_letter + str(offset_stat_row + 12)] = "=" + str(batches[batch_index]) + " * " + col_letter + str(offset_stat_row + 7)
        inference_sheet[col_letter + str(offset_stat_row + 13)] = "=" + str(batches[batch_index]) + " * " + col_letter + str(offset_stat_row + 8)
        inference_sheet[col_letter + str(offset_stat_row + 14)] = "=" + str(batches[batch_index]) + " * " + col_letter + str(offset_stat_row + 9)
        inference_sheet[col_letter + str(offset_stat_row + 15)] = "=" + str(batches[batch_index]) + " * " + col_letter + str(offset_stat_row + 10)
        inference_sheet[col_letter + str(offset_stat_row + 16)] = "=" + str(batches[batch_index]) + " * " + col_letter + str(offset_stat_row + 11)
        inference_sheet[col_letter + str(offset_stat_row + 17)] = str(warm_up_times[batches[batch_index]])

    chart = LineChart()
    chart.title = "Metrics"
    chart.x_axis.title = "Batch Size"
    chart.y_axis.title = "Time (s)"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = x_axis_max
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    metrics = ["Average", "Median", "90th Percentile", "95th Percentile", "99th Percentile", "Minimum", "Maximum"]
    for metric_index in range(len(metrics)):
        series = Series(values=Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row + metric_index, max_col=offset_col + len(batches) - 1, max_row=offset_stat_row + metric_index), title=f"{metrics[metric_index]}")
        series.marker.symbol = "circle"
        series.marker.size = 6
        chart.series.append(series)
    batch_titles = Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row - 1, max_col=offset_col + len(batches) - 1, max_row=offset_stat_row - 1)
    chart.set_categories(batch_titles)
    chart.legend.position = 'b'
    chart.layout=Layout(
        manualLayout=ManualLayout(
            x=0.02, y=0.02,
            h=0.65, w=0.9,
        )
    )
    chart.width = 25
    inference_sheet.add_chart(chart, get_column_letter(len(batches) + 2) + "1")
    main_sheet.add_chart(deepcopy(chart), "F5")

    chart = LineChart()
    chart.title = "IPS"
    chart.x_axis.title = "Batch Size"
    chart.y_axis.title = "Inferences Per Second"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    metrics = ["Average", "Median", "90th Percentile", "95th Percentile", "99th Percentile"]
    for metric_index in range(len(metrics)):
        series = Series(values=Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row + 7 + metric_index, max_col=offset_col + len(batches) - 1, max_row=offset_stat_row + 7 + metric_index), title=f"{metrics[metric_index]}")
        series.marker.symbol = "circle"
        series.marker.size = 6
        chart.series.append(series)
    batch_titles = Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row - 1, max_col=offset_col + len(batches) - 1, max_row=offset_stat_row - 1)
    chart.set_categories(batch_titles)
    chart.legend.position = 'b'
    chart.layout=Layout(
        manualLayout=ManualLayout(
            x=0.02, y=0.02,
            h=0.65, w=0.9,
        )
    )
    chart.width = 15
    inference_sheet.add_chart(chart, get_column_letter(len(batches) + 2) + "16")
    main_sheet.add_chart(deepcopy(chart), "F20")

    chart = LineChart()
    chart.title = "BPS"
    chart.x_axis.title = "Batch Size"
    chart.y_axis.title = "Batches Per Second"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    metrics = ["Average", "Median", "90th Percentile", "95th Percentile", "99th Percentile"]
    for metric_index in range(len(metrics)):
        series = Series(values=Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row + 12 + metric_index, max_col=offset_col + len(batches) - 1, max_row=offset_stat_row + 12 + metric_index), title=f"{metrics[metric_index]}")
        series.marker.symbol = "circle"
        series.marker.size = 6
        chart.series.append(series)
    batch_titles = Reference(inference_sheet, min_col=offset_col, min_row=offset_stat_row - 1, max_col=offset_col + len(batches) - 1, max_row=offset_stat_row - 1)
    chart.set_categories(batch_titles)
    chart.legend.position = 'b'
    chart.layout=Layout(
        manualLayout=ManualLayout(
            x=0.02, y=0.02,
            h=0.65, w=0.9,
        )
    )
    chart.width = 15
    inference_sheet.add_chart(chart, get_column_letter(len(batches) + 11) + "16")
    main_sheet.add_chart(deepcopy(chart), "P20")

    idx = 0
    for idx in range(1, len(inference_table[0]) + 1):
        row = [idx]
        for batch_index in range(len(batches)):
            row.append(inference_table[batch_index][idx - 1] if len(inference_table[batch_index]) > idx - 1 else None)
        inference_sheet.append(row)
    chart = LineChart()
    chart.title = "Inference times"
    chart.x_axis.title = "Run"
    chart.y_axis.title = "Time (s)"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = x_axis_max
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    for batch_index in range(len(batches)):
        series = Series(values=Reference(inference_sheet, min_col=batch_index + offset_col, min_row=offset_row, max_col=batch_index + offset_col, max_row=last_row), title=f"Batch {batches[batch_index]}")
        chart.series.append(series)
    chart.width = 15
    chart.legend.position = 'b'
    chart.layout=Layout(
        manualLayout=ManualLayout(
            x=0.02, y=0.02,
            h=0.65, w=0.9,
        )
    )
    inference_sheet.add_chart(chart, get_column_letter(len(batches) + 2) + "33")
    main_sheet.add_chart(deepcopy(chart), "P35")

    report_datetime = datetime.datetime.now()
    main_sheet.title = "Overview"
    main_sheet.column_dimensions[get_column_letter(1)].width = 30
    main_sheet.append(['Model:', model_name])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)
    main_sheet.append(['Description:', str(model)])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)
    main_sheet.append(['Run Command:', ' '.join(sys.argv)])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)
    main_sheet.append(['Report Date:', report_datetime.strftime('%Y-%m-%d %H:%M:%S')])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=6)
    main_sheet.append(['Batches:', *batches])
    main_sheet.append(['Total Inference Runs:', model.total_inference_runs])
    main_sheet.append([])
    main_sheet.append(['System Information:'])
    try:
        main_sheet.append(['Hostname:', platform.node()])
        main_sheet.append(['OS:', platform.system()])
        main_sheet.append(['OS Version:', platform.version()])
        main_sheet.append(['OS Release:', platform.release()])
    except Exception as e:
        main_sheet.append([f'Cannot get OS information {e}'])
    main_sheet.append(['Python Version:', sys.version])
    main_sheet.merge_cells(start_row=main_sheet.max_row, start_column=2, end_row=main_sheet.max_row, end_column=10)

    try:
        main_sheet.append(['CPU:', platform.processor()])
        accelerators = enumerate_accelerators()
        for item in accelerators['gpu']:
            main_sheet.append(['GPU:', item['name']])
        for item in accelerators['npu']:
            main_sheet.append(['NPU:', item['name']])
    except Exception as e:
        main_sheet.append([f'Cannot get accelerators information {e}'])

    try:
        result = subprocess.run(
            ['pip', 'list', '--format', 'columns'],
            capture_output=True,
            text=True
        )
        output = result.stdout.split('\n')
        for item in output:
            main_sheet.append(item.split())
    except Exception as e:
        main_sheet.append([f'Cannot get PIP list {e}'])

    try:
        main_sheet.append(['Loaded Modules:'])
        for item in sorted(list_loaded_modules()['modules'], key=lambda x: x['name']):
            main_sheet.append([item['name'], item['path']])
    except Exception as e:
        main_sheet.append([f'Cannot get loaded modules {e}'])

    try:
        main_sheet.append([])
        main_sheet.append(['Environment Variables:'])
        for key, value in sorted(os.environ.items(), key=lambda x: x[0]):
            main_sheet.append([key, value])
    except Exception as e:
        main_sheet.append([f'Cannot get environment variables {e}'])

    workbook_path = f"{platform.node().lower()}_{model_name}_{report_datetime.strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(workbook_path)

    reports_path = os.path.join(os.path.dirname(__file__), 'reports', report_datetime.strftime("%Y%m%d"))
    if not os.path.exists(reports_path):
      os.makedirs(reports_path)
      # Copying statistics aggregator to a reports folder
      try:
        copy(os.path.join(os.path.dirname(__file__), "!StatViewer.xlsm"), os.path.join(reports_path, "!StatViewer.xlsm"))
      except Exception as e:
        print(f'{{ "Error": "Failed to copy !StatViewer.xlsm {e}" }}')
    os.rename(workbook_path, os.path.join(reports_path, workbook_path))

    print(f"{{ \"Workbook\": \"{os.path.join(reports_path, workbook_path).replace('\\', '/')}\" }},")

  except Exception as e:
    print(f'{{ "Error": "Failed to load openpyxl {e}" }},')
  return workbook_path

def _run(cmd, timeout=5):
    try:
        p = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout, check=False)
        out = (p.stdout or "").strip()
        err = (p.stderr or "").strip()
        return out, err, p.returncode
    except Exception as e:
        return "", str(e), 1

def _windows_gpus():
    # Uses built-in PowerShell + CIM (WMI) to enumerate GPUs
    ps = [
        "powershell", "-NoProfile", "-Command",
        r"Get-CimInstance Win32_VideoController | "
        r"Select-Object Name,AdapterCompatibility,PNPDeviceID | ConvertTo-Json -Depth 3"
    ]
    out, _, rc = _run(ps, timeout=8)
    if rc != 0 or not out:
        return []
    try:
        import json
        data = json.loads(out)
        if isinstance(data, dict):
            data = [data]
        return [
            {
                "name": d.get("Name"),
                "vendor": d.get("AdapterCompatibility"),
                "pnp_device_id": d.get("PNPDeviceID"),
            }
            for d in data
            if d.get("Name")
        ]
    except Exception:
        return []

def _windows_npus():
    # Windows 11 often exposes NPUs under PnP class "Neural"
    ps = [
        "powershell", "-NoProfile", "-Command",
        r"Get-CimInstance Win32_PnPEntity | "
        r"Where-Object { $_.PNPClass -eq 'Neural' -or $_.Name -match '(?i)\bNPU\b|Neural Engine|Neural Processing|AI Accelerator' } | "
        r"Select-Object Name,PNPClass,DeviceID | ConvertTo-Json -Depth 3"
    ]
    out, _, rc = _run(ps, timeout=10)
    if rc != 0 or not out:
        return []
    try:
        import json
        data = json.loads(out)
        if isinstance(data, dict):
            data = [data]
        return [
            {"name": d.get("Name"), "class": d.get("PNPClass"), "device_id": d.get("DeviceID")}
            for d in data
            if d.get("Name")
        ]
    except Exception:
        return []

def _linux_lspci_lines():
    if not which("lspci"):
        return []
    out, _, rc = _run(["lspci", "-nn"], timeout=5)
    if rc != 0 or not out:
        return []
    return out.splitlines()

def _linux_gpus():
    lines = _linux_lspci_lines()
    gpu_markers = ("VGA compatible controller", "3D controller", "Display controller")
    gpus = []
    for ln in lines:
        if any(m in ln for m in gpu_markers):
            gpus.append({"name": ln})
    return gpus

def _linux_npus():
    lines = _linux_lspci_lines()
    # PCI class "Processing accelerators" is common for AI/NPUs, plus keyword heuristics
    npus = []
    import re
    for ln in lines:
        if ("Processing accelerators" in ln) or re.search(r"(?i)\bNPU\b|Neural|AI accelerator|TPU", ln):
            npus.append({"name": ln})
    return npus

def _mac_gpus():
    if not which("system_profiler"):
        return []
    out, _, rc = _run(["system_profiler", "SPDisplaysDataType", "-json"], timeout=10)
    if rc != 0 or not out:
        return []
    try:
        import json
        data = json.loads(out)
        items = data.get("SPDisplaysDataType", [])
        gpus = []
        for it in items:
            # Keys vary by macOS version; keep it simple
            name = it.get("sppci_model") or it.get("_name")
            if name:
                gpus.append({"name": name, "raw": it})
        return gpus
    except Exception:
        return []

def _mac_npus():
    # Apple Neural Engine shows up in hardware profile text on Apple Silicon
    if not which("system_profiler"):
        return []
    out, _, rc = _run(["system_profiler", "SPHardwareDataType"], timeout=8)
    if rc != 0 or not out:
        return []
    import re
    m = re.search(r"Neural Engine:\s*(.+)", out)
    return [{"name": f"Apple Neural Engine ({m.group(1).strip()})"}] if m else []

def enumerate_accelerators():
    osname = platform.system()
    if osname == "Windows":
        return {"gpu": _windows_gpus(), "npu": _windows_npus()}
    if osname == "Linux":
        return {"gpu": _linux_gpus(), "npu": _linux_npus()}
    if osname == "Darwin":
        return {"gpu": _mac_gpus(), "npu": _mac_npus()}
    return {"gpu": [], "npu": []}

def list_loaded_modules():
    osname = platform.system()
    result = {
        'pid': os.getpid(),
        'executable': sys.executable,
        'modules': []
    }

    if osname == "Windows":
        result['modules'] = _windows_list_modules()
    elif osname == "Linux":
        result['modules'] = _linux_list_modules()
    elif osname == "Darwin":
        result['modules'] = _mac_list_modules()

    return result

def _windows_list_modules():
    modules = []

    # Try using psutil first (most reliable cross-platform method)
    try:
        import psutil
        process = psutil.Process()
        for dll in process.memory_maps():
            modules.append({
                'name': os.path.basename(dll.path),
                'path': dll.path
            })
        return modules
    except ImportError:
        pass
    except Exception:
        pass

    # Fallback to ctypes approach
    try:
        import ctypes
        from ctypes import wintypes

        kernel32 = ctypes.windll.kernel32
        psapi = ctypes.windll.psapi

        hProcess = kernel32.GetCurrentProcess()

        hMods = (wintypes.HMODULE * 1024)()
        cbNeeded = wintypes.DWORD()

        if psapi.EnumProcessModules(hProcess, ctypes.byref(hMods), ctypes.sizeof(hMods), ctypes.byref(cbNeeded)):
            count = int(cbNeeded.value / ctypes.sizeof(wintypes.HMODULE))

            for i in range(count):
                module_name = ctypes.create_unicode_buffer(260)
                module_path = ctypes.create_unicode_buffer(260)

                if psapi.GetModuleFileNameExW(hProcess, hMods[i], module_path, ctypes.sizeof(module_path)):
                    if psapi.GetModuleBaseNameW(hProcess, hMods[i], module_name, ctypes.sizeof(module_name)):
                        modules.append({
                            'name': module_name.value,
                            'path': module_path.value,
                            'base_address': hex(hMods[i]) if hMods[i] else None
                        })
    except Exception as e:
        # If all else fails, return error info
        modules.append({'error': str(e)})

    return modules

def _linux_list_modules():
    modules = []
    seen_paths = set()

    # Try using psutil first
    try:
        import psutil
        process = psutil.Process()
        for mmap in process.memory_maps():
            if mmap.path and mmap.path not in seen_paths:
                seen_paths.add(mmap.path)
                modules.append({
                    'name': os.path.basename(mmap.path),
                    'path': mmap.path
                })
        return modules
    except ImportError:
        pass
    except Exception:
        pass

    # Fallback to reading /proc/self/maps
    try:
        with open('/proc/self/maps', 'r') as f:
            for line in f:
                parts = line.split()
                if len(parts) >= 6:
                    pathname = ' '.join(parts[5:])
                    if pathname and pathname not in ['[stack]', '[heap]', '[vdso]', '[vsyscall]']:
                        if pathname.startswith('/') and pathname not in seen_paths:
                            seen_paths.add(pathname)
                            address = parts[0].split('-')[0]
                            modules.append({
                                'name': os.path.basename(pathname),
                                'path': pathname,
                                'base_address': '0x' + address
                            })
    except Exception as e:
        modules.append({'error': str(e)})

    return modules

def _mac_list_modules():
    modules = []

    # Try using psutil first
    try:
        import psutil
        process = psutil.Process()
        for mmap in process.memory_maps():
            if mmap.path:
                modules.append({
                    'name': os.path.basename(mmap.path),
                    'path': mmap.path
                })
        return modules
    except ImportError:
        pass
    except Exception:
        pass

    # Fallback to vmmap command
    try:
        pid = os.getpid()
        out, _, rc = _run(['vmmap', str(pid)], timeout=10)
        if rc == 0 and out:
            seen_paths = set()
            for line in out.splitlines():
                if '/' in line:
                    parts = line.split()
                    for part in parts:
                        if part.startswith('/') and os.path.exists(part):
                            if part not in seen_paths:
                                seen_paths.add(part)
                                modules.append({
                                    'name': os.path.basename(part),
                                    'path': part
                                })
    except Exception as e:
        modules.append({'error': str(e)})

    return modules
