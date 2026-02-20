import os
import sys
import random
import datetime
import pytest
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from reports import performance_report


class _MockModel:
    """Minimal stand-in for a model object required by performance_report."""
    def __init__(self, total_inference_runs=100):
        self.total_inference_runs = total_inference_runs

    def __str__(self):
        return "MockModel for testing"

class TestPerformanceReport:

    @pytest.fixture(autouse=True)
    def setup(self, tmp_path, monkeypatch):
        self.parent_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.reports_base = os.path.join(self.parent_dir, 'reports')
        monkeypatch.chdir(tmp_path)
        self._cleanup_paths = []
        yield
        for p in self._cleanup_paths:
            try:
                if os.path.exists(p):
                    os.remove(p)
            except OSError:
                pass

    def _resolve_workbook(self, workbook_path):
        """Return absolute path where performance_report moved the file."""
        date_str = datetime.datetime.now().strftime("%Y%m%d")
        full = os.path.join(self.reports_base, date_str, workbook_path)
        assert os.path.isfile(full), f"Workbook not found at {full}"
        self._cleanup_paths.append(full)
        return full

    def _make_data(self, batches, n_read=10, n_inference=12, seed=42):
        random.seed(seed)
        # last element is the summary
        read_times = [random.uniform(0.01, 1.0) for _ in range(n_read)] + [{"Minimum": 0.01, "Maximum": 1.0, "Average": 0.5}]
        inference_times = {}
        warm_up_times = {}
        for b in batches:
            inference_times[b] = [random.uniform(0.01, 1.0) for _ in range(n_inference)] + [{"Minimum": 0.01, "Maximum": 1.0, "Average": 0.5}]
            warm_up_times[b] = random.uniform(0.5, 5.0)
        return read_times, inference_times, warm_up_times

    def _generate_report(self, batches, n_read=10, n_inference=12, seed=42):
        read_times, inference_times, warm_up_times = self._make_data(
            batches, n_read, n_inference, seed
        )
        model = _MockModel(total_inference_runs=10)
        model_name = f"test_{datetime.datetime.now().strftime('%H%M%S%f')}"
        workbook_path = performance_report(
            model, model_name, read_times, inference_times, warm_up_times, batches
        )
        full_path = self._resolve_workbook(workbook_path)
        return full_path, read_times, inference_times, warm_up_times

    def test_report_worksheets_and_overview(self):
        """Generate a report, verify worksheet names and Overview labels."""
        full_path, *_ = self._generate_report(
            batches=[1], n_read=3, n_inference=4
        )

        wb = openpyxl.load_workbook(full_path)

        expected_sheets = ["Overview", "Read", "Inference"]
        for name in expected_sheets:
            assert name in wb.sheetnames, f"Missing worksheet '{name}'"

        overview = wb["Overview"]
        labels = set()
        for row in overview.iter_rows(min_col=1, max_col=1, values_only=True):
            if row[0] is not None:
                labels.add(str(row[0]))

        checkable_labels = [
            "Model:",
            "Description:",
            "Run Command:",
            "Report Date:",
            "Batches:",
            "Total Inference Runs:",
            "System Information:",
            "Hostname:",
            "OS:",
            "OS Version:",
            "OS Release:",
            "Python Version:",
            "CPU:",
        ]
        for lbl in checkable_labels:
            assert lbl in labels, f"Expected '{lbl}' on Overview sheet"

        assert "Loaded Modules:" in labels, \
            "Expected 'Loaded Modules:' on Overview sheet"
        assert "Environment Variables:" in labels, \
            "Expected 'Environment Variables:' on Overview sheet"

        wb.close()

    @pytest.mark.parametrize("batches", [[1], [2, 4, 8]])
    def test_data_values(self, batches):
        """Verify stored cell values match the data passed to performance_report."""
        n_read, n_inference = 10, 12
        full_path, read_times, inference_times, warm_up_times = \
            self._generate_report(batches, n_read=n_read, n_inference=n_inference)
        
        wb = openpyxl.load_workbook(full_path)

        self._verify_read_sheet(wb["Read"], read_times)
        self._verify_inference_sheet(
            wb["Inference"], inference_times, warm_up_times, batches
        )

        wb.close()

    @staticmethod
    def _find_row(sheet, col_a_value, col_b_value=None):
        """Row number (1-based) where column A matches, optionally column B too."""
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row[0] == col_a_value:
                if col_b_value is None or (len(row) > 1 and row[1] == col_b_value):
                    return idx
        return None

    def _verify_read_sheet(self, sheet, read_times):
        assert isinstance(read_times[-1], dict), "Last element of read_times is not a dictionary"
        assert "Minimum" in read_times[-1], "Last element of read_times is missing 'Minimum'"
        assert "Maximum" in read_times[-1], "Last element of read_times is missing 'Maximum'"
        assert "Average" in read_times[-1], "Last element of read_times is missing 'Average'"
        # performance_report stores read_times[:-1]
        expected = read_times[:-1]

        run_row = self._find_row(sheet, "Run", "Time (s)")
        assert run_row is not None, "Read sheet: missing 'Run / Time (s)' header"
        stat_rows = {
            'Average': '=AVERAGE({col}{row_s}:{col}{row_e})',
            'Median': '=MEDIAN({col}{row_s}:{col}{row_e})',
            '90th Percentile': '=_xlfn.PERCENTILE.INC({col}{row_s}:{col}{row_e}, 0.9)',
            '95th Percentile': '=_xlfn.PERCENTILE.INC({col}{row_s}:{col}{row_e}, 0.95)',
            '99th Percentile': '=_xlfn.PERCENTILE.INC({col}{row_s}:{col}{row_e}, 0.99)',
            'Minimum': '=MIN({col}{row_s}:{col}{row_e})',
            'Maximum': '=MAX({col}{row_s}:{col}{row_e})',
        }

        actual = []
        data_rows = run_row + 1
        for row in sheet.iter_rows(
            min_row=run_row + 1, min_col=2, max_col=2, values_only=True
        ):
            if row[0] is not None:
                data_rows += 1
                actual.append(row[0])

        assert len(actual) == len(expected), \
            f"Read data rows: expected {len(expected)}, got {len(actual)}"
        for i, (a, e) in enumerate(zip(actual, expected)):
            assert a == pytest.approx(e), \
                f"Read row {i + 1}: expected {e}, got {a}"
        for name in stat_rows.keys():
            row = self._find_row(sheet, name)
            assert row is not None, f"Inference sheet: missing '{name}' row"
            col = openpyxl.utils.get_column_letter(2)
            row_s = run_row + 1
            row_e = data_rows - 1 # it points on row after last data row
            cell_value = sheet.cell(row=row, column=2).value
            expected_value = stat_rows[name].format(col=col, row_s=row_s, row_e=row_e)
            assert str(cell_value) == str(expected_value), f"Read sheet: '{name}' row value is incorrect, expected {expected_value}, got {cell_value}"

    def _verify_inference_sheet(self, sheet, inference_times, warm_up_times, batches):
        wu_row = self._find_row(sheet, "Warm Up Time")
        assert wu_row is not None, "Inference sheet: missing 'Warm Up Time' row"

        for bi, batch in enumerate(batches):
            stored = sheet.cell(row=wu_row, column=2 + bi).value
            assert stored == str(warm_up_times[batch]), \
                f"Warm-up batch {batch}: expected '{warm_up_times[batch]}', got '{stored}'"

        run_row = self._find_row(sheet, "Run")
        assert run_row is not None, "Inference sheet: missing 'Run' header"
        data_rows = {}

        stat_rows = {
            'Average': '=AVERAGE({col}{row_s}:{col}{row_e})',
            'Median': '=MEDIAN({col}{row_s}:{col}{row_e})',
            '90th Percentile': '=_xlfn.PERCENTILE.INC({col}{row_s}:{col}{row_e}, 0.9)',
            '95th Percentile': '=_xlfn.PERCENTILE.INC({col}{row_s}:{col}{row_e}, 0.95)',
            '99th Percentile': '=_xlfn.PERCENTILE.INC({col}{row_s}:{col}{row_e}, 0.99)',
            'Minimum': '=MIN({col}{row_s}:{col}{row_e})',
            'Maximum': '=MAX({col}{row_s}:{col}{row_e})',
            'IPS (Average)': '=1 / {col}3',
            'IPS (Median)': '=1 / {col}4',
            'IPS (90th Percentile)': '=1 / {col}5',
            'IPS (95th Percentile)': '=1 / {col}6',
            'IPS (99th Percentile)': '=1 / {col}7',
            'BPS (Average)': '={batch} * {col}10',
            'BPS (Median)': '={batch} * {col}11',
            'BPS (90th Percentile)': '={batch} * {col}12',
            'BPS (95th Percentile)': '={batch} * {col}13',
            'BPS (99th Percentile)': '={batch} * {col}14',
        }

        for bi, batch in enumerate(batches):
            assert isinstance(inference_times[batch][-1], dict), "Last element of inference_times is not a dictionary"
            assert "Minimum" in inference_times[batch][-1], "Last element of inference_times is missing 'Minimum'"
            assert "Maximum" in inference_times[batch][-1], "Last element of inference_times is missing 'Maximum'"
            assert "Average" in inference_times[batch][-1], "Last element of inference_times is missing 'Average'"

            table = inference_times[batch][:-1]
            expected = table

            data_rows[batch] = run_row + 1
            actual = []
            for row in sheet.iter_rows(
                min_row=run_row + 1, min_col=2 + bi, max_col=2 + bi,
                values_only=True
            ):
                if row[0] is not None:
                    actual.append(row[0])
                    data_rows[batch] += 1

            assert len(actual) == len(expected), \
                f"Inference batch {batch}: expected {len(expected)} rows, got {len(actual)}"
            for i, (a, e) in enumerate(zip(actual, expected)):
                assert a == pytest.approx(e), \
                    f"Inference batch {batch} row {i + 1}: expected {e}, got {a}"

            for name in stat_rows.keys():
                row = self._find_row(sheet, name)
                assert row is not None, f"Inference sheet: missing '{name}' row"
                col = openpyxl.utils.get_column_letter(2 + bi)
                row_s = run_row + 1
                row_e = data_rows[batch] - 1 # it points on row after last data row
                cell_value = sheet.cell(row=row, column=2 + bi).value
                expected_value = stat_rows[name].format(col=col, row_s=row_s, row_e=row_e, batch=batch)
                assert str(cell_value) == str(expected_value), f"Inference sheet: '{name}' row value is incorrect for batch {batch}, expected {expected_value}, got {cell_value}"
